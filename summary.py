"""
summary.py — Attendance Summary Generator (FIXED)
Run this AFTER main.py to see who is Present/Absent today.
Usage: py -3.11 summary.py
"""

import os
import csv
import datetime

# ─────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Folder where person subfolders (face images) are stored
FACES_DIR = os.path.join(BASE_DIR, "Attendance_data")

# All possible folders where attendance CSV might be saved
POSSIBLE_CSV_FOLDERS = [
    os.path.join(BASE_DIR, "Attendance_Entry"),
    os.path.join(BASE_DIR, "Attendance_data"),
    os.path.join(BASE_DIR, "attendance_data"),
    os.path.join(BASE_DIR, "attendance"),
    BASE_DIR,
]

def normalize(name):
    """Normalize name for comparison — strip spaces, uppercase."""
    return name.strip().upper()


def get_all_registered_names():
    """Get all registered person names from subfolders."""
    names = []
    if not os.path.exists(FACES_DIR):
        print(f"[ERROR] Faces folder not found: {FACES_DIR}")
        return names
    for item in os.listdir(FACES_DIR):
        item_path = os.path.join(FACES_DIR, item)
        if os.path.isdir(item_path) and not item.startswith('.'):
            names.append(item.strip())
    return sorted(names)


def find_todays_csv():
    """Search all possible folders for today's attendance CSV."""
    today  = datetime.datetime.now().strftime("%Y_%m_%d")
    today2 = datetime.datetime.now().strftime("%Y-%m-%d")
    today3 = datetime.datetime.now().strftime("%d-%m-%Y")

    patterns = [
        f"Attendance_{today}.csv",
        f"Attendance_{today2}.csv",
        f"attendance_{today}.csv",
        f"Attendance_{today3}.csv",
    ]

    for folder in POSSIBLE_CSV_FOLDERS:
        if not os.path.exists(folder):
            continue
        for pattern in patterns:
            path = os.path.join(folder, pattern)
            if os.path.exists(path):
                print(f"[INFO] Found attendance file: {path}")
                return path
        try:
            for f in os.listdir(folder):
                if f.endswith(".csv") and "Summary" not in f and (today in f or today2 in f):
                    full = os.path.join(folder, f)
                    print(f"[INFO] Found attendance file: {full}")
                    return full
        except:
            pass

    print(f"[WARNING] No attendance CSV found for today ({today}).")
    print("          Please run main.py first!\n")
    return None


def get_present_names(csv_path):
    """Read CSV and return set of normalized present names."""
    present = set()
    if not csv_path:
        return present

    with open(csv_path, newline='', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        print(f"[INFO] CSV columns: {headers}")

        # Find name column
        name_col = None
        for h in headers:
            if h.strip().lower() in ['name', 'student_name', 'person', 'student']:
                name_col = h
                break
        if not name_col and headers:
            name_col = headers[0]
            print(f"[INFO] Using first column as name: '{name_col}'")

        if not name_col:
            print("[ERROR] Could not find name column!")
            return present

        f.seek(0)
        reader = csv.DictReader(f)
        for row in reader:
            raw = row.get(name_col, "").strip()
            if raw:
                present.add(normalize(raw))

    print(f"[INFO] Names found in CSV: {present}")
    return present


def print_summary(all_names, present_norm):
    today = datetime.datetime.now().strftime("%d-%m-%Y")
    total = len(all_names)
    present_count = sum(1 for n in all_names if normalize(n) in present_norm)
    absent_count  = total - present_count

    print("\n")
    print("=" * 55)
    print(f"        ATTENDANCE SUMMARY — {today}")
    print("=" * 55)
    print(f"  {'#':<5} {'NAME':<28} {'STATUS'}")
    print("-" * 55)
    for i, name in enumerate(all_names, 1):
        status = "✅ PRESENT" if normalize(name) in present_norm else "❌ ABSENT"
        print(f"  {i:<5} {name:<28} {status}")
    print("-" * 55)
    print(f"  Total Registered : {total}")
    print(f"  Present Today    : {present_count}")
    print(f"  Absent Today     : {absent_count}")
    print("=" * 55)
    print()


def save_summary_excel(all_names, present_norm):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError:
        os.system("py -3.11 -m pip install openpyxl")
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    today_display = datetime.datetime.now().strftime("%d-%m-%Y")
    today_file    = datetime.datetime.now().strftime("%Y_%m_%d")
    excel_file    = os.path.join(FACES_DIR, f"Summary_{today_file}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Summary {today_display}"

    header_fill  = PatternFill("solid", fgColor="1B4F72")
    present_fill = PatternFill("solid", fgColor="D5F5E3")
    absent_fill  = PatternFill("solid", fgColor="FADBD8")
    title_fill   = PatternFill("solid", fgColor="0D1B2A")
    footer_fill  = PatternFill("solid", fgColor="D6EAF8")
    header_font  = Font(bold=True, color="FFFFFF", size=12)
    present_font = Font(bold=True, color="1E8449", size=11)
    absent_font  = Font(bold=True, color="C0392B", size=11)
    normal_font  = Font(size=11)
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")
    thin   = Border(
        left=Side(style="thin", color="BDC3C7"),
        right=Side(style="thin", color="BDC3C7"),
        top=Side(style="thin", color="BDC3C7"),
        bottom=Side(style="thin", color="BDC3C7"),
    )

    # Title row
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value     = f"ATTENDANCE SUMMARY — {today_display}"
    c.font      = Font(bold=True, color="FFFFFF", size=14)
    c.fill      = title_fill
    c.alignment = center
    ws.row_dimensions[1].height = 30

    # Header row
    for col, h in enumerate(["S.No", "Name", "Status", "Date"], 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = center; cell.border = thin
    ws.row_dimensions[2].height = 22

    # Data rows
    for i, name in enumerate(all_names, 1):
        is_present  = normalize(name) in present_norm
        status_text = "PRESENT" if is_present else "ABSENT"
        rfill = present_fill if is_present else absent_fill
        sfont = present_font if is_present else absent_font
        r = i + 2
        for col, val in enumerate([i, name, status_text, today_display], 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.fill = rfill; cell.border = thin
            cell.alignment = left if col == 2 else center
            cell.font = sfont if col == 3 else normal_font
        ws.row_dimensions[r].height = 20

    # Footer
    total   = len(all_names)
    present = sum(1 for n in all_names if normalize(n) in present_norm)
    absent  = total - present
    fr = total + 3
    ws.merge_cells(f"A{fr}:D{fr}")
    fc = ws[f"A{fr}"]
    fc.value     = f"Total: {total}   |   Present: {present}   |   Absent: {absent}"
    fc.font      = Font(bold=True, size=11, color="1B2631")
    fc.fill      = footer_fill
    fc.alignment = center
    ws.row_dimensions[fr].height = 22

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16

    wb.save(excel_file)
    print(f"[INFO] ✅ Excel saved: {excel_file}")
    return excel_file


# ─────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("\n[INFO] Reading registered persons...")
    all_names = get_all_registered_names()

    if not all_names:
        print("[ERROR] No registered persons found!")
        exit(1)

    print(f"[INFO] Found {len(all_names)} persons: {all_names}\n")

    print("[INFO] Searching for today's attendance file...")
    csv_path = find_todays_csv()

    print("[INFO] Reading present students...")
    present_norm = get_present_names(csv_path)
    print(f"[INFO] Present: {len(present_norm)} students\n")

    print_summary(all_names, present_norm)
    save_summary_excel(all_names, present_norm)

    print("[INFO] Done! Open Summary Excel in Attendance_data folder.")
    print("       Green = PRESENT   |   Red = ABSENT\n")