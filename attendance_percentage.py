"""
attendance_percent.py — Monthly Attendance Percentage Report
Compiles ALL past attendance data and gives percentage for each student till date.
Run this daily after summary.py.
Usage: py -3.11 attendance_percent.py
"""

import os
import csv
import datetime

# ─────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

FACES_DIR = os.path.join(BASE_DIR, "Attendance_data")

POSSIBLE_CSV_FOLDERS = [
    os.path.join(BASE_DIR, "Attendance_Entry"),
    os.path.join(BASE_DIR, "Attendance_data"),
    os.path.join(BASE_DIR, "attendance_data"),
    os.path.join(BASE_DIR, "attendance"),
    BASE_DIR,
]

LOW_PERCENT_THRESHOLD = 75  # Below this % → highlighted in red


def normalize(name):
    return name.strip().upper()


def get_all_registered_names():
    names = []
    if not os.path.exists(FACES_DIR):
        print(f"[ERROR] Faces folder not found: {FACES_DIR}")
        return names
    for item in os.listdir(FACES_DIR):
        item_path = os.path.join(FACES_DIR, item)
        if os.path.isdir(item_path) and not item.startswith('.'):
            names.append(item.strip())
    return sorted(names)


def get_all_attendance_csvs():
    """Find ALL past attendance CSV files (not summary files)."""
    all_csvs = []
    for folder in POSSIBLE_CSV_FOLDERS:
        if not os.path.exists(folder):
            continue
        try:
            for f in os.listdir(folder):
                if (f.endswith(".csv")
                        and "Summary" not in f
                        and "Percent" not in f
                        and f.lower().startswith("attendance")):
                    full_path = os.path.join(folder, f)
                    if full_path not in all_csvs:
                        all_csvs.append(full_path)
        except:
            pass
    all_csvs.sort()
    return all_csvs


def extract_date_from_filename(filename):
    """Try to extract a date from CSV filename."""
    import re
    base = os.path.basename(filename).replace(".csv", "")
    # Match patterns like 2026_04_21 or 2026-04-21 or 21-04-2026
    patterns = [
        r'(\d{4})[_\-](\d{2})[_\-](\d{2})',  # YYYY_MM_DD
        r'(\d{2})[_\-](\d{2})[_\-](\d{4})',  # DD_MM_YYYY
    ]
    for p in patterns:
        m = re.search(p, base)
        if m:
            g = m.groups()
            try:
                if len(g[0]) == 4:  # YYYY first
                    return datetime.date(int(g[0]), int(g[1]), int(g[2]))
                else:               # DD first
                    return datetime.date(int(g[2]), int(g[1]), int(g[0]))
            except:
                pass
    return None


def read_present_names_from_csv(csv_path):
    """Return set of normalized names who were present in this CSV."""
    present = set()
    try:
        with open(csv_path, newline='', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames or []
            name_col = None
            for h in headers:
                if h.strip().lower() in ['name', 'student_name', 'person', 'student']:
                    name_col = h
                    break
            if not name_col and headers:
                name_col = headers[0]
            if not name_col:
                return present
            f.seek(0)
            reader = csv.DictReader(f)
            for row in reader:
                raw = row.get(name_col, "").strip()
                if raw:
                    present.add(normalize(raw))
    except Exception as e:
        print(f"[WARNING] Could not read {csv_path}: {e}")
    return present


def compile_attendance_data(all_names, all_csvs):
    """
    Returns:
      - per_student: dict { norm_name: { date: True/False, ... } }
      - all_dates: sorted list of date objects
      - monthly: dict { norm_name: { 'YYYY-MM': (present_days, total_days) } }
    """
    norm_names = {normalize(n): n for n in all_names}
    per_student = {normalize(n): {} for n in all_names}
    all_dates = []

    for csv_path in all_csvs:
        date = extract_date_from_filename(csv_path)
        if date is None:
            continue
        if date in all_dates:
            continue
        all_dates.append(date)
        present = read_present_names_from_csv(csv_path)
        for norm in per_student:
            per_student[norm][date] = (norm in present)

    all_dates.sort()

    # Build monthly summary
    monthly = {normalize(n): {} for n in all_names}
    for norm, date_dict in per_student.items():
        for date, was_present in date_dict.items():
            month_key = date.strftime("%Y-%m")
            if month_key not in monthly[norm]:
                monthly[norm][month_key] = [0, 0]  # [present, total]
            monthly[norm][month_key][1] += 1
            if was_present:
                monthly[norm][month_key][0] += 1

    return per_student, all_dates, monthly


def print_percent_summary(all_names, per_student, all_dates):
    today_str = datetime.date.today().strftime("%d-%m-%Y")
    total_days = len(all_dates)

    print("\n")
    print("=" * 65)
    print(f"      ATTENDANCE PERCENTAGE REPORT — Till {today_str}")
    print("=" * 65)
    print(f"  {'#':<4} {'NAME':<26} {'PRESENT':<10} {'TOTAL':<8} {'%'}")
    print("-" * 65)

    for i, name in enumerate(all_names, 1):
        norm = normalize(name)
        present_count = sum(1 for v in per_student[norm].values() if v)
        pct = (present_count / total_days * 100) if total_days > 0 else 0
        flag = " ⚠ LOW" if pct < LOW_PERCENT_THRESHOLD else ""
        print(f"  {i:<4} {name:<26} {present_count:<10} {total_days:<8} {pct:.1f}%{flag}")

    print("-" * 65)
    print(f"  Total working days recorded: {total_days}")
    print("=" * 65)
    print()


def save_percent_excel(all_names, per_student, all_dates, monthly):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError:
        os.system("py -3.11 -m pip install openpyxl")
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    today_str  = datetime.date.today().strftime("%d-%m-%Y")
    today_file = datetime.date.today().strftime("%Y_%m_%d")
    excel_file = os.path.join(FACES_DIR, f"Attendance_Percentage_{today_file}.xlsx")

    wb = openpyxl.Workbook()

    # ── Colors & styles ────────────────────────────────────────
    def fill(color): return PatternFill("solid", fgColor=color)
    def font(bold=False, color="000000", size=11):
        return Font(bold=bold, color=color, size=size)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center")
    def border():
        s = Side(style="thin", color="BDC3C7")
        return Border(left=s, right=s, top=s, bottom=s)

    # ═══════════════════════════════════════════════════════════
    # SHEET 1 — Overall Summary (till date)
    # ═══════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Overall Summary"
    total_days = len(all_dates)

    # Title
    ws1.merge_cells("A1:F1")
    c = ws1["A1"]
    c.value     = f"ATTENDANCE PERCENTAGE — Till {today_str}"
    c.font      = Font(bold=True, color="FFFFFF", size=14)
    c.fill      = fill("0D1B2A")
    c.alignment = center
    ws1.row_dimensions[1].height = 30

    # Sub-header
    ws1.merge_cells("A2:F2")
    c2 = ws1["A2"]
    c2.value     = f"Total Working Days: {total_days}"
    c2.font      = Font(bold=True, color="FFFFFF", size=11)
    c2.fill      = fill("1B4F72")
    c2.alignment = center
    ws1.row_dimensions[2].height = 20

    # Column headers
    headers = ["S.No", "Student Name", "Days Present", "Days Absent", "Total Days", "Attendance %"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=3, column=col, value=h)
        cell.fill = fill("2E86C1")
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = center
        cell.border = border()
    ws1.row_dimensions[3].height = 22

    # Data rows
    for i, name in enumerate(all_names, 1):
        norm          = normalize(name)
        present_count = sum(1 for v in per_student[norm].values() if v)
        absent_count  = total_days - present_count
        pct           = (present_count / total_days * 100) if total_days > 0 else 0
        r = i + 3

        is_low = pct < LOW_PERCENT_THRESHOLD
        row_bg = fill("FADBD8") if is_low else (fill("D5F5E3") if pct >= 90 else fill("FEF9E7"))

        data = [i, name, present_count, absent_count, total_days, f"{pct:.1f}%"]
        for col, val in enumerate(data, 1):
            cell = ws1.cell(row=r, column=col, value=val)
            cell.fill      = row_bg
            cell.border    = border()
            cell.alignment = left if col == 2 else center
            if col == 6:
                cell.font = Font(bold=True, size=11,
                                 color="C0392B" if is_low else ("1E8449" if pct >= 90 else "D68910"))
            else:
                cell.font = Font(size=11)
        ws1.row_dimensions[r].height = 20

    # Legend
    leg_row = len(all_names) + 5
    ws1.merge_cells(f"A{leg_row}:F{leg_row}")
    lc = ws1[f"A{leg_row}"]
    lc.value     = "Green >= 90% (Good)   |   Yellow = 75-89% (Average)   |   Red < 75% (LOW ATTENDANCE)"
    lc.font      = Font(bold=True, size=10, color="1B2631")
    lc.fill      = fill("D6EAF8")
    lc.alignment = center
    ws1.row_dimensions[leg_row].height = 20

    # Column widths
    ws1.column_dimensions["A"].width = 7
    ws1.column_dimensions["B"].width = 28
    ws1.column_dimensions["C"].width = 14
    ws1.column_dimensions["D"].width = 14
    ws1.column_dimensions["E"].width = 13
    ws1.column_dimensions["F"].width = 15

    # ═══════════════════════════════════════════════════════════
    # SHEET 2 — Month-wise Breakdown
    # ═══════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Month-wise")

    # Get all unique months
    all_months = sorted(set(
        date.strftime("%Y-%m")
        for date in all_dates
    ))
    month_labels = [datetime.datetime.strptime(m, "%Y-%m").strftime("%b %Y") for m in all_months]

    # Title
    total_cols = 2 + len(all_months) * 2 + 1
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c = ws2.cell(row=1, column=1, value=f"MONTH-WISE ATTENDANCE — Till {today_str}")
    c.font      = Font(bold=True, color="FFFFFF", size=14)
    c.fill      = fill("0D1B2A")
    c.alignment = center
    ws2.row_dimensions[1].height = 30

    # Month headers (merged across Present/%)
    col = 3
    for m_label in month_labels:
        ws2.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+1)
        c = ws2.cell(row=2, column=col, value=m_label)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = fill("1B4F72")
        c.alignment = center
        c.border = border()
        col += 2

    # Overall % header
    ws2.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
    oc = ws2.cell(row=2, column=col, value="Overall %")
    oc.font = Font(bold=True, color="FFFFFF", size=11)
    oc.fill = fill("0D1B2A")
    oc.alignment = center
    oc.border = border()

    # Sub-headers: S.No, Name, then Present/% per month
    ws2.cell(row=3, column=1, value="S.No").font  = Font(bold=True, color="FFFFFF", size=10)
    ws2.cell(row=3, column=1).fill = fill("2E86C1")
    ws2.cell(row=3, column=1).alignment = center
    ws2.cell(row=3, column=1).border = border()
    ws2.cell(row=3, column=2, value="Student Name").font = Font(bold=True, color="FFFFFF", size=10)
    ws2.cell(row=3, column=2).fill = fill("2E86C1")
    ws2.cell(row=3, column=2).alignment = center
    ws2.cell(row=3, column=2).border = border()

    col = 3
    for _ in all_months:
        for label in ["Present", "%"]:
            c = ws2.cell(row=3, column=col, value=label)
            c.font = Font(bold=True, color="FFFFFF", size=10)
            c.fill = fill("2E86C1")
            c.alignment = center
            c.border = border()
            col += 1

    ws2.row_dimensions[2].height = 22
    ws2.row_dimensions[3].height = 20

    # Data rows
    for i, name in enumerate(all_names, 1):
        norm  = normalize(name)
        r = i + 3

        # Overall
        total_present = sum(1 for v in per_student[norm].values() if v)
        overall_pct   = (total_present / total_days * 100) if total_days > 0 else 0
        is_low_overall = overall_pct < LOW_PERCENT_THRESHOLD

        ws2.cell(row=r, column=1, value=i).alignment = center
        ws2.cell(row=r, column=1).border = border()
        ws2.cell(row=r, column=1).font = Font(size=10)

        ws2.cell(row=r, column=2, value=name).alignment = left
        ws2.cell(row=r, column=2).border = border()
        ws2.cell(row=r, column=2).font = Font(size=10)

        col = 3
        for m_key in all_months:
            m_data = monthly[norm].get(m_key, [0, 0])
            m_present, m_total = m_data
            m_pct = (m_present / m_total * 100) if m_total > 0 else 0
            is_low = m_pct < LOW_PERCENT_THRESHOLD and m_total > 0

            bg = fill("FADBD8") if is_low else (fill("D5F5E3") if m_pct >= 90 else fill("FEF9E7"))

            pc = ws2.cell(row=r, column=col, value=m_present if m_total > 0 else "-")
            pc.fill = bg; pc.alignment = center; pc.border = border(); pc.font = Font(size=10)

            pctc = ws2.cell(row=r, column=col+1,
                            value=f"{m_pct:.0f}%" if m_total > 0 else "-")
            pctc.fill = bg; pctc.alignment = center; pctc.border = border()
            pctc.font = Font(bold=True, size=10,
                             color="C0392B" if is_low else ("1E8449" if m_pct >= 90 else "D68910"))
            col += 2

        # Overall % cell
        oc = ws2.cell(row=r, column=col, value=f"{overall_pct:.1f}%")
        oc.fill = fill("FADBD8") if is_low_overall else (fill("D5F5E3") if overall_pct >= 90 else fill("FEF9E7"))
        oc.alignment = center; oc.border = border()
        oc.font = Font(bold=True, size=10,
                       color="C0392B" if is_low_overall else ("1E8449" if overall_pct >= 90 else "D68910"))

        ws2.row_dimensions[r].height = 20

    # Column widths for sheet 2
    ws2.column_dimensions["A"].width = 7
    ws2.column_dimensions["B"].width = 26
    col_letter = 3
    for _ in all_months:
        for offset in range(2):
            from openpyxl.utils import get_column_letter
            ws2.column_dimensions[get_column_letter(col_letter + offset)].width = 10
        col_letter += 2
    from openpyxl.utils import get_column_letter
    ws2.column_dimensions[get_column_letter(col_letter)].width = 12

    # ═══════════════════════════════════════════════════════════
    # SHEET 3 — Day-by-Day Detail
    # ═══════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Day-by-Day")

    date_labels = [d.strftime("%d-%m") for d in all_dates]
    total_cols3 = 2 + len(all_dates) + 1

    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols3)
    c = ws3.cell(row=1, column=1, value=f"DAY-BY-DAY ATTENDANCE — Till {today_str}")
    c.font = Font(bold=True, color="FFFFFF", size=14)
    c.fill = fill("0D1B2A"); c.alignment = center
    ws3.row_dimensions[1].height = 30

    ws3.cell(row=2, column=1, value="S.No").font = Font(bold=True, color="FFFFFF", size=10)
    ws3.cell(row=2, column=1).fill = fill("2E86C1"); ws3.cell(row=2, column=1).alignment = center
    ws3.cell(row=2, column=1).border = border()
    ws3.cell(row=2, column=2, value="Name").font = Font(bold=True, color="FFFFFF", size=10)
    ws3.cell(row=2, column=2).fill = fill("2E86C1"); ws3.cell(row=2, column=2).alignment = center
    ws3.cell(row=2, column=2).border = border()

    for j, dl in enumerate(date_labels):
        c = ws3.cell(row=2, column=3+j, value=dl)
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.fill = fill("1B4F72"); c.alignment = center; c.border = border()

    ws3.cell(row=2, column=3+len(all_dates), value="Total %").font = Font(bold=True, color="FFFFFF", size=10)
    ws3.cell(row=2, column=3+len(all_dates)).fill = fill("0D1B2A")
    ws3.cell(row=2, column=3+len(all_dates)).alignment = center
    ws3.cell(row=2, column=3+len(all_dates)).border = border()
    ws3.row_dimensions[2].height = 22

    for i, name in enumerate(all_names, 1):
        norm = normalize(name)
        r = i + 2
        ws3.cell(row=r, column=1, value=i).alignment = center
        ws3.cell(row=r, column=1).border = border(); ws3.cell(row=r, column=1).font = Font(size=9)
        ws3.cell(row=r, column=2, value=name).alignment = left
        ws3.cell(row=r, column=2).border = border(); ws3.cell(row=r, column=2).font = Font(size=9)

        for j, date in enumerate(all_dates):
            was_present = per_student[norm].get(date, False)
            c = ws3.cell(row=r, column=3+j, value="P" if was_present else "A")
            c.fill = fill("D5F5E3") if was_present else fill("FADBD8")
            c.alignment = center; c.border = border()
            c.font = Font(bold=True, size=9,
                          color="1E8449" if was_present else "C0392B")

        # Total %
        present_count = sum(1 for v in per_student[norm].values() if v)
        pct = (present_count / total_days * 100) if total_days > 0 else 0
        is_low = pct < LOW_PERCENT_THRESHOLD
        tc = ws3.cell(row=r, column=3+len(all_dates), value=f"{pct:.1f}%")
        tc.fill = fill("FADBD8") if is_low else (fill("D5F5E3") if pct >= 90 else fill("FEF9E7"))
        tc.alignment = center; tc.border = border()
        tc.font = Font(bold=True, size=9,
                       color="C0392B" if is_low else ("1E8449" if pct >= 90 else "D68910"))
        ws3.row_dimensions[r].height = 18

    ws3.column_dimensions["A"].width = 6
    ws3.column_dimensions["B"].width = 24
    for j in range(len(all_dates)):
        from openpyxl.utils import get_column_letter
        ws3.column_dimensions[get_column_letter(3+j)].width = 6
    from openpyxl.utils import get_column_letter
    ws3.column_dimensions[get_column_letter(3+len(all_dates))].width = 10

    wb.save(excel_file)
    print(f"[INFO] Percentage report saved: {excel_file}")
    return excel_file


# ─────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("\n[INFO] Reading registered persons...")
    all_names = get_all_registered_names()
    if not all_names:
        print("[ERROR] No registered persons found!")
        exit(1)
    print(f"[INFO] Found {len(all_names)} persons: {all_names}\n")

    print("[INFO] Scanning all past attendance CSV files...")
    all_csvs = get_all_attendance_csvs()
    if not all_csvs:
        print("[ERROR] No attendance CSV files found! Run main.py first.")
        exit(1)
    print(f"[INFO] Found {len(all_csvs)} attendance file(s).\n")

    print("[INFO] Compiling attendance data...")
    per_student, all_dates, monthly = compile_attendance_data(all_names, all_csvs)

    print_percent_summary(all_names, per_student, all_dates)
    save_percent_excel(all_names, per_student, all_dates, monthly)

    print("[INFO] Done! Open Attendance_Percentage Excel in Attendance_data folder.")
    print("       Sheet 1 = Overall Summary")
    print("       Sheet 2 = Month-wise Breakdown")
    print("       Sheet 3 = Day-by-Day Detail\n")